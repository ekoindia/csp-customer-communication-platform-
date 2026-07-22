"""
In-memory Excel envelope for centralized OCR results.

The server turns OCR rows into an .xlsx entirely in RAM (`rows_to_xlsx_bytes`);
the CSP turns that .xlsx back into rows entirely in RAM (`xlsx_bytes_to_rows`).
Nothing is ever written to disk on either side — the bytes only live inside the
encrypted OCR envelope in transit and in memory while being built/parsed.

Why an .xlsx round trip at all (vs. just shipping the rows as JSON): the CSP
already has a battle-tested Excel ingestion path (parser -> column_mapper ->
validator -> review gate) used for real bank Excel uploads. Feeding server OCR
output through that SAME path means one downstream code path to maintain and
test, not two, and the server output behaves exactly like a bank Excel file.

CRITICAL: every cell is written as TEXT. Account numbers and mobile numbers can
have leading zeros; a numeric Excel cell would silently drop them ("09876..." ->
9876...). Writing/reading everything as strings makes the round trip lossless.
"""
import io
from typing import Dict, List

# Stable, human-readable column order. These header names are chosen so the
# CSP-side column_mapper maps them straight back to the same canonical fields
# (it matches on substrings: "account", "name", "mobile", "balance"/"band").
# Any extra keys present on a row are appended after these, in first-seen order.
_PREFERRED_COLUMNS = [
    "account_number", "name", "mobile", "balance_band",
    "father_name", "village", "taluka", "address",
]

_SHEET_TITLE = "ocr"


def _ordered_columns(rows: List[Dict]) -> List[str]:
    """Only columns that actually appear in the data, preferred order first.

    We deliberately do NOT emit preferred columns that no row carries — that
    would write empty father_name/address columns the MVP never populates
    (and, for PII hygiene, we don't want empty PII columns in the sheet)."""
    present = set()
    for row in rows:
        present.update(row.keys())
    cols = [c for c in _PREFERRED_COLUMNS if c in present]
    seen = set(cols)
    for row in rows:
        for key in row.keys():
            if key not in seen:
                cols.append(key)
                seen.add(key)
    return cols


def rows_to_xlsx_bytes(rows: List[Dict]) -> bytes:
    """Serialize OCR rows to .xlsx bytes in memory. All cells are text."""
    import openpyxl

    rows = rows or []
    columns = _ordered_columns(rows)
    # write_only keeps memory flat for large batches (no full in-RAM cell tree).
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title=_SHEET_TITLE)
    ws.append(columns)
    for row in rows:
        ws.append(["" if row.get(c) is None else str(row.get(c)) for c in columns])
    buf = io.BytesIO()
    try:
        wb.save(buf)
    finally:
        wb.close()
    return buf.getvalue()


def _normalise_header(value, index: int) -> str:
    if value is None:
        return f"col_{index}"
    return str(value).strip().lower().replace(" ", "_")


def xlsx_bytes_to_rows(data: bytes) -> List[Dict]:
    """Parse .xlsx bytes (from the server) back into row dicts, in memory.

    Mirrors core.parser._parse_excel exactly (same header normalisation, same
    "skip fully-empty rows" rule) but reads from a bytes buffer instead of a
    path, so the file never touches the CSP's disk. Values stay as strings."""
    import openpyxl

    if not data:
        return []
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        ws = wb.active
        raw = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not raw:
        return []
    headers = [_normalise_header(h, i) for i, h in enumerate(raw[0])]
    out: List[Dict] = []
    for row in raw[1:]:
        if not any(cell is not None and str(cell) != "" for cell in row):
            continue
        out.append({h: ("" if v is None else str(v)) for h, v in zip(headers, row)})
    return out
